import openpyxl
from scripts.ingestion.models import AbstractIngestor, UnsupportedFileError

class XlsxIngestor(AbstractIngestor):
    """
    Ingestor for XLSX files. Extracts sheet-wise content.
    """

    def ingest(self, filepath: str) -> list[tuple[str, dict]]:
        if not filepath.endswith(".xlsx"):
            raise UnsupportedFileError("File is not a .xlsx file.")

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            extracted_data = []

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines = []

                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    line = "\t".join(str(cell) for cell in row if cell is not None)
                    if line.strip():
                        lines.append(line)

                sheet_text = "\n".join(lines).strip()
                if sheet_text:
                    metadata = {
                        "sheet_name": sheet,
                        "type": "sheet_content",
                        "doc_type": "xlsx"
                    }
                    extracted_data.append((sheet_text, metadata))

            return extracted_data

        except Exception as e:
            raise UnsupportedFileError(f"Failed to process XLSX file: {filepath}: {e}")
